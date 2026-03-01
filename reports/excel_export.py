# ==============================================================================
# KatilimALM - (c) 2024-2026 Salih Say
# GitHub: github.com/SalihSay
# Bu yazilim telif haklari ile korunmaktadir.
# Izinsiz kopyalanmasi, dagitilmasi veya degistirilmesi yasaktir.
# ==============================================================================
"""
KatılımALM — Excel Rapor Yardımcıları
Formatlı, düzenli Excel çıktıları üretir.
Otomatik sütun genişliği, para birimi formatı, renkli başlıklar.
"""
from io import BytesIO
from typing import List
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))


# Stiller
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical='center', wrap_text=False)
MONEY_FORMAT = '#,##0'
MONEY_FORMAT_DECIMAL = '#,##0.00'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)
SUCCESS_FILL = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
DANGER_FILL = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
ACCENT_FILL = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')


def auto_fit_columns(ws, min_width=10, max_width=45):
    """Sütun genişliklerini içeriğe göre otomatik ayarla."""
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def style_header_row(ws):
    """Başlık satırını formatla."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30


def style_data_rows(ws, money_columns=None, percent_columns=None):
    """
    Veri satırlarını formatla.
    money_columns: Para formatı uygulanacak sütun indeksleri (1-indexed)
    percent_columns: Yüzde formatı uygulanacak sütun indeksleri
    """
    if money_columns is None:
        money_columns = []
    if percent_columns is None:
        percent_columns = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        for col_idx, cell in enumerate(row, 1):
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGNMENT
            
            # Alternatif satır renklendirme
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            
            # Para formatı
            if col_idx in money_columns and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT
            
            # Yüzde formatı
            if col_idx in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'


def style_status_column(ws, col_idx, green_values=None, red_values=None):
    """Durum sütununu renklendir."""
    if green_values is None:
        green_values = ['Uygun', 'Evet', '✅']
    if red_values is None:
        red_values = ['Aşım', 'Hayır', '❌', 'Uyarı']
    
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row:
            val = str(cell.value) if cell.value else ''
            if val in green_values:
                cell.fill = SUCCESS_FILL
            elif val in red_values:
                cell.fill = DANGER_FILL


def format_money_with_currency(amount, currency='TL'):
    """Para tutarını para birimi ile birlikte formatla."""
    if amount is None:
        return ''
    return f"{amount:,.0f} {currency}"


def create_formatted_excel(
    balance_sheet, off_balance, metrics, bank_name, report_date, report_type
):
    """
    Tam formatlı Excel rapor üretir.
    Her sayfada: başlık formatı, otomatik genişlik, para birimi, renkler.
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # ====================================================================
        # 1. KAPAK SAYFASI
        # ====================================================================
        kapak = pd.DataFrame({
            'Bilgi': [
                'Banka Adı', 'Rapor Tarihi', 'Rapor Tipi', 
                'Oluşturan Sistem', 'Gizlilik Derecesi', '',
                'Toplam Aktif', 'Toplam Pasif', 'Özkaynak'
            ],
            'Değer': [
                bank_name, 
                report_date.strftime('%d.%m.%Y'), 
                report_type,
                'KatılımALM v1.0', 
                '— GİZLİ —',
                '',
                format_money_with_currency(sum(i.amount for i in balance_sheet if i.side == 'aktif')),
                format_money_with_currency(sum(i.amount for i in balance_sheet if i.side == 'pasif')),
                format_money_with_currency(sum(i.amount for i in balance_sheet if 'ozkaynak' in i.instrument_type)),
            ],
        })
        kapak.to_excel(writer, sheet_name='Kapak', index=False)
        ws = writer.sheets['Kapak']
        style_header_row(ws)
        auto_fit_columns(ws, min_width=18, max_width=50)
        # Gizlilik satırını kırmızı yap
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT
        
        # ====================================================================
        # 2. YÖNETİCİ ÖZETİ
        # ====================================================================
        ozet = pd.DataFrame({
            'Gösterge': [
                'Likidite Karşılama Oranı (LCR)', 
                'Net Kararlı Fonlama Oranı (NSFR)',
                'Kaldıraç Oranı',
                'Vade Uyumsuzluğu (Duration Gap)',
                'Aktif Taraf Ortalama Vade',
                'Pasif Taraf Ortalama Vade',
            ],
            'Değer': [
                f"%{metrics['lcr'].lcr_ratio:.1f}",
                f"%{metrics['nsfr'].nsfr_ratio:.1f}",
                f"%{metrics['leverage'].leverage_ratio:.1f}",
                f"{metrics['dur_gap']:.2f} yıl",
                f"{metrics['a_dur']:.2f} yıl",
                f"{metrics['p_dur']:.2f} yıl",
            ],
            'BDDK Limiti': ['Min %100', 'Min %100', 'Min %3', '—', '—', '—'],
            'Durum': [
                'Uygun' if metrics['lcr'].is_compliant else 'Aşım',
                'Uygun' if metrics['nsfr'].is_compliant else 'Aşım',
                'Uygun' if metrics['leverage'].is_compliant else 'Aşım',
                '—', '—', '—',
            ],
        })
        ozet.to_excel(writer, sheet_name='Yönetici Özeti', index=False)
        ws = writer.sheets['Yönetici Özeti']
        style_header_row(ws)
        style_data_rows(ws)
        style_status_column(ws, 4)
        auto_fit_columns(ws, min_width=15, max_width=40)
        
        # ====================================================================
        # 3. BİLANÇO
        # ====================================================================
        bs_data = []
        for item in balance_sheet:
            currency_label = {
                'TL': 'Türk Lirası (TL)', 'USD': 'ABD Doları (USD)', 
                'EUR': 'Euro (EUR)', 'XAU': 'Altın (XAU)'
            }.get(item.currency, item.currency)
            
            bs_data.append({
                'Kalem Adı': item.name,
                'Tutar (TL Karşılığı)': round(item.amount),
                'Para Birimi': currency_label,
                'Orijinal Tutar': round(item.amount_original) if item.amount_original else '',
                'Aktif / Pasif': 'AKTİF' if item.side == 'aktif' else 'PASİF',
                'Enstrüman Türü (Açıklama)': _get_friendly_instrument_name(item.instrument_type),
                'İslami Finans Sınıfı': item.islamic_class,
                'Kalan Vade (Gün)': item.maturity_days if item.maturity_days < 99999 else 'Süresiz',
                'Yıllık Kâr Payı Oranı (%)': round(item.profit_rate * 100, 2) if item.profit_rate > 0 else '—',
                'TMSF Sigortalı': 'Evet' if item.is_insured else '—',
                'Karşı Taraf Türü': _get_friendly_counterparty(item.counterparty_type),
            })
        
        df_bs = pd.DataFrame(bs_data)
        df_bs.to_excel(writer, sheet_name='Bilanço', index=False)
        ws = writer.sheets['Bilanço']
        style_header_row(ws)
        style_data_rows(ws, money_columns=[2, 4])
        auto_fit_columns(ws, min_width=12, max_width=35)
        
        # ====================================================================
        # 4. LCR NAKİT ÇIKIŞLARI
        # ====================================================================
        lcr_out = metrics['lcr'].outflow_detail
        if lcr_out:
            out_data = []
            for item in lcr_out:
                out_data.append({
                    'Kalem Adı': item['name'],
                    'Tutar (TL)': round(item['amount']),
                    'Kaçış Oranı (Run-off)': f"%{item['runoff_rate']*100:.0f}",
                    '30 Günlük Nakit Çıkışı (TL)': round(item['outflow']),
                    'Kaynak': 'Bilanço İçi' if item.get('source') == 'bilanço_içi' else 'Bilanço Dışı',
                    'Para Birimi': item.get('currency', 'TL'),
                })
            df_out = pd.DataFrame(out_data)
            df_out.to_excel(writer, sheet_name='LCR Nakit Çıkışları', index=False)
            ws = writer.sheets['LCR Nakit Çıkışları']
            style_header_row(ws)
            style_data_rows(ws, money_columns=[2, 4])
            auto_fit_columns(ws, min_width=12, max_width=40)
        
        # ====================================================================
        # 5. LCR NAKİT GİRİŞLERİ
        # ====================================================================
        lcr_in = metrics['lcr'].inflow_detail
        if lcr_in:
            in_data = []
            for item in lcr_in:
                in_data.append({
                    'Kalem Adı': item['name'],
                    'Tutar (TL)': round(item['amount']),
                    'Giriş Oranı': f"%{item['inflow_rate']*100:.0f}",
                    '30 Günlük Nakit Girişi (TL)': round(item['inflow']),
                    'Para Birimi': item.get('currency', 'TL'),
                })
            df_in = pd.DataFrame(in_data)
            df_in.to_excel(writer, sheet_name='LCR Nakit Girişleri', index=False)
            ws = writer.sheets['LCR Nakit Girişleri']
            style_header_row(ws)
            style_data_rows(ws, money_columns=[2, 4])
            auto_fit_columns(ws)
        
        # ====================================================================
        # 6. NSFR — KULLANILABILIR KARARLI FONLAMA (ASF)
        # ====================================================================
        if metrics['nsfr'].asf.items_detail:
            asf_data = []
            for item in metrics['nsfr'].asf.items_detail:
                asf_data.append({
                    'Kalem Adı': item['name'],
                    'Tutar (TL)': round(item['amount']),
                    'ASF Ağırlığı (%)': f"%{item['weight']*100:.0f}",
                    'ASF Katkısı (TL)': round(item['contribution']),
                    'Para Birimi': item.get('currency', 'TL'),
                    'Kalan Vade (Gün)': item.get('maturity_days', '—'),
                })
            df_asf = pd.DataFrame(asf_data)
            df_asf.to_excel(writer, sheet_name='NSFR Kararlı Fonlama (ASF)', index=False)
            ws = writer.sheets['NSFR Kararlı Fonlama (ASF)']
            style_header_row(ws)
            style_data_rows(ws, money_columns=[2, 4])
            auto_fit_columns(ws)
        
        # ====================================================================
        # 7. NSFR — GEREKLİ KARARLI FONLAMA (RSF)
        # ====================================================================
        if metrics['nsfr'].rsf.items_detail:
            rsf_data = []
            for item in metrics['nsfr'].rsf.items_detail:
                rsf_data.append({
                    'Kalem Adı': item['name'],
                    'Tutar (TL)': round(item['amount']),
                    'RSF Ağırlığı (%)': f"%{item['weight']*100:.0f}",
                    'RSF Gereksinimi (TL)': round(item['contribution']),
                    'Para Birimi': item.get('currency', 'TL'),
                    'Kaynak': item.get('source', '—'),
                })
            df_rsf = pd.DataFrame(rsf_data)
            df_rsf.to_excel(writer, sheet_name='NSFR Gerekli Fonlama (RSF)', index=False)
            ws = writer.sheets['NSFR Gerekli Fonlama (RSF)']
            style_header_row(ws)
            style_data_rows(ws, money_columns=[2, 4])
            auto_fit_columns(ws)
        
        # ====================================================================
        # 8. VADE UYUMSUZLUĞU (GAP ANALİZİ)
        # ====================================================================
        gap_data = []
        for g in metrics['gap_table']:
            gap_data.append({
                'Vade Aralığı': g.bucket_name,
                'Faize Duyarlı Varlıklar (TL)': round(g.rate_sensitive_assets),
                'Faize Duyarlı Yükümlülükler (TL)': round(g.rate_sensitive_liabilities),
                'Açık/Fazla (Gap) (TL)': round(g.gap),
                'Birikimli Açık (TL)': round(g.cumulative_gap),
                'Gap / Toplam Aktif (%)': round(g.gap_to_total_assets, 2),
            })
        df_gap = pd.DataFrame(gap_data)
        df_gap.to_excel(writer, sheet_name='Vade Uyumsuzluğu (Gap)', index=False)
        ws = writer.sheets['Vade Uyumsuzluğu (Gap)']
        style_header_row(ws)
        style_data_rows(ws, money_columns=[2, 3, 4, 5])
        auto_fit_columns(ws, min_width=14, max_width=35)
        
        # ====================================================================
        # 9. STRES TESTİ SONUÇLARI
        # ====================================================================
        stress_data = []
        for r in metrics['stress']:
            stress_data.append({
                'Senaryo Adı': r.scenario.name,
                'Açıklama': r.scenario.description,
                'Kur Şoku (TL Değer Kaybı %)': f"%{r.scenario.fx_shock*100:.0f}",
                'Kâr Payı Şoku (Baz Puan)': f"{r.scenario.rate_shock_bp} bp",
                'Mevduat Kaçışı (%)': f"%{r.scenario.deposit_runoff*100:.0f}",
                'Stres Sonrası LCR (%)': f"%{r.stressed_lcr:.1f}",
                'Stres Sonrası NSFR (%)': f"%{r.stressed_nsfr:.1f}",
                'LCR Etkisi': f"{r.lcr_impact:+.1f}",
                'LCR Uygunluk': 'Uygun' if r.lcr_compliant else 'Aşım',
                'NSFR Uygunluk': 'Uygun' if r.nsfr_compliant else 'Aşım',
            })
        df_stress = pd.DataFrame(stress_data)
        df_stress.to_excel(writer, sheet_name='Stres Testi', index=False)
        ws = writer.sheets['Stres Testi']
        style_header_row(ws)
        style_data_rows(ws)
        style_status_column(ws, 9)
        style_status_column(ws, 10)
        auto_fit_columns(ws, min_width=12, max_width=30)
        
        # ====================================================================
        # 10. KÂR PAYI HAVUZLARI
        # ====================================================================
        if metrics['pools'].pools:
            pool_data = []
            for p in metrics['pools'].pools:
                pool_data.append({
                    'Havuz Adı': p.pool_name,
                    'Toplam Fon (TL)': round(p.total_funds),
                    'Kullandırılan Tutar (TL)': round(p.total_placements),
                    'Fon Kullanım Oranı (%)': f"%{p.fund_utilization:.1f}",
                    'Aylık Net Kâr (TL)': round(p.net_income),
                    'Banka Payı (Alpha)': f"%{p.bank_share_ratio*100:.0f}",
                    'Müşteriye Dağıtılan Kâr Payı Oranı (Yıllık %)': f"%{p.profit_rate*100:.2f}",
                })
            df_pools = pd.DataFrame(pool_data)
            df_pools.to_excel(writer, sheet_name='Kâr Payı Havuzları', index=False)
            ws = writer.sheets['Kâr Payı Havuzları']
            style_header_row(ws)
            style_data_rows(ws, money_columns=[2, 3, 5])
            auto_fit_columns(ws, min_width=14, max_width=40)
    
    output.seek(0)
    return output


def _get_friendly_instrument_name(instrument_type):
    """Teknik enstrüman adını anlaşılır Türkçeye çevirir."""
    names = {
        'nakit': 'Nakit ve Kasa',
        'merkez_bankasi': 'Merkez Bankası Hesabı',
        'devlet_sukuk': 'Devlet Sukuk (Kira Sertifikası)',
        'devlet_kira_sertifikasi': 'Devlet Kira Sertifikası',
        'ozel_sukuk_aa': 'Özel Sektör Sukuk (Yüksek Derece)',
        'ozel_sukuk_diger': 'Özel Sektör Sukuk (Standart)',
        'murabaha_alacak': 'Murabaha Finansmanı (Maliyet + Kâr)',
        'finansal_kiralama': 'Finansal Kiralama (İcara)',
        'bankalararasi_murabaha': 'Bankalar Arası Plasman',
        'bankalararasi_borc': 'Bankalar Arası Borçlanma',
        'sabit_varlik': 'Sabit Varlıklar (Bina, Ekipman)',
        'diger_aktif': 'Diğer Varlıklar',
        'katilma_vadesiz': 'Vadesiz Hesap (Cari Hesap)',
        'katilma_vadeli': 'Vadeli Katılma Hesabı (Kâr Payı)',
        'ihrac_sukuk': 'Banka Tarafından İhraç Edilen Sukuk',
        'ozkaynaklar': 'Özkaynak (Sermaye + Yedekler)',
        'diger_yukumluluk': 'Diğer Borçlar',
    }
    return names.get(instrument_type, instrument_type)


def _get_friendly_counterparty(counterparty):
    """Karşı taraf türünü anlaşılır Türkçeye çevirir."""
    names = {
        'perakende': 'Bireysel Müşteri',
        'kurumsal': 'Kurumsal Müşteri',
        'finansal': 'Banka / Finans Kuruluşu',
        'devlet': 'Devlet / Hazine',
        'retail': 'Bireysel Müşteri',
        'corporate': 'Kurumsal Müşteri',
        '': '—',
    }
    return names.get(counterparty, counterparty if counterparty else '—')
